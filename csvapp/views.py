import openpyxl

from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect
from django.db import IntegrityError, transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser

from .serializers import FuneralPolicySerializer, IndluLoanDataSerializer, SmartAdvanceCreditSerializer

from .models import FuneralPolicy, IndluLoanData, SmartAdvanceCredit
from .forms import CSVFuneralPolicyUpload, CSVIndluLoanDataUpload, CSVSmartAdvanceCredit

class FuneralPolicyList(APIView):

    parser_classes = [MultiPartParser]

    def get(self, request, format=None):
        funeral_policies = FuneralPolicy.objects.all()
        serializer = FuneralPolicySerializer(funeral_policies, many=True)
        return Response(serializer.data)
    
    def post(self, request, format=None):

        if 'csv_file' not in request.FILES:
            return Response({'error': 'Please provide CSV file'}, status=status.HTTP_400_BAD_REQUEST)

        csv_file = request.FILES['csv_file']

        try:
            wb = openpyxl.load_workbook(csv_file)
            worksheet = wb["Funeral Policies"]

            new_records = []
            updated_records = []
            failed_records = []

            for row in worksheet.iter_rows(values_only=True):

                policy_number = row[10]

                try:
                    # Updates existing record

                    funeral_policy = FuneralPolicy.objects.get(policy_number=policy_number)
                    for i, field in enumerate(FuneralPolicy._meta.get_fields()):
                        setattr(funeral_policy, field.name, row[i])

                    funeral_policy.save()
                    updated_records.append({"policy_number": policy_number, "status": "updated"})

                except FuneralPolicy.DoesNotExist:
                    
                    funeral_policy = FuneralPolicy(
                        policy_search=row[0],
                        ID_search=row[1],
                        time_stamp=row[2],
                        report_period_start=row[3],
                        report_period_end=row[4],
                        administrator=row[5],
                        insurer_name=row[6],
                        client_identifier=row[7],
                        division_identifier=row[8],
                        sub_scheme_name=row[9],
                        policy_number=row[10],
                        product_name=row[11],
                        product_option=row[12],
                        policy_commencement_date=row[13],
                        policy_expiry_date=row[14],
                        term_of_policy=row[15],
                        policy_status=row[16],
                        policy_status_date=row[17],
                        new_policy_indicator=row[18],
                        sales_channel=row[19],
                        cancelled_by=row[20],
                        death_indicator=row[21],
                        premium_frequency=row[22],
                        premium_type=row[23],
                        death_original_sum_insured=row[24],
                        death_cover=row[25],
                        death_current_sum_insured=row[26],
                        reinsurer_name=row[27],
                        death_current_ri_sum_insured=row[28],
                        death_ri_premium=row[29],
                        death_ri_percent=row[30],
                        total_premium_collected=row[31],
                        total_premium_payable=row[32],
                        total_premium_subsidy=row[33],
                        total_reinsurance_premium=row[34],
                        total_reinsurance_premium_payable=row[35],
                        total_financial_reinsurance_cashflows=row[36],
                        total_financial_reinsurance_payable=row[37],
                        commission_frequency=row[38],
                        commission=row[39],
                        binder_fees=row[40],
                        outsourcing_fees=row[41],
                        marketing_fees=row[42],
                        management_fees=row[43],
                        claims_handling_fees=row[44],
                        total_gross_claim_amount=row[45],
                        gross_claim_paid=row[46],
                        reinsurance_recoveries=row[47],
                        principal_surname=row[48],
                        principal_firstname=row[49],
                        principal_initials=row[50],
                        principal_id=row[51],
                        principal_gender=row[52],
                        principal_date_of_birth=row[53],
                        principal_physical_address=row[54],
                        principal_postal_code=row[55],
                        principal_email=row[56],
                        income_group=row[57],
                        spouse_indicator=row[58],
                        number_adult_dependants=row[59],
                        number_child_dependants=row[60],
                        number_extended_family=row[61],
                        spouse_surname=row[62],
                        spouse_firstname=row[63],
                        spouse_initials=row[64],
                        spouse_id=row[65],
                        spouse_gender=row[66],
                        spouse_date_of_birth=row[67],
                        spouse_cover_amount=row[68],
                        spouse_cover_commencement_date=row[69],

                        dependent_1_gender=row[70],
                        dependent_1_date_of_birth=row[71],
                        dependent_1_type=row[72],
                        dependent_1_cover_amount=row[73],
                        dependent_1_cover_commencement_date=row[74],

                        dependent_2_gender=row[75],
                        dependent_2_date_of_birth=row[76],
                        dependent_2_type=row[77],
                        dependent_2_cover_amount=row[78],
                        dependent_2_cover_commencement_date=row[79],

                        dependent_3_gender=row[80],
                        dependent_3_date_of_birth=row[81],
                        dependent_3_type=row[82],
                        dependent_3_cover_amount=row[83],
                        dependent_3_cover_commencement_date=row[84],

                        dependent_4_gender=row[85],
                        dependent_4_date_of_birth=row[86],
                        dependent_4_type=row[87],
                        dependent_4_cover_amount=row[88],
                        dependent_4_cover_commencement_date=row[89],

                        dependent_5_gender=row[90],
                        dependent_5_date_of_birth=row[91],
                        dependent_5_type=row[92],
                        dependent_5_cover_amount=row[93],
                        dependent_5_cover_commencement_date=row[94],

                        dependent_6_gender=row[95],
                        dependent_6_date_of_birth=row[96],
                        dependent_6_type=row[97],
                        dependent_6_cover_amount=row[98],
                        dependent_6_cover_commencement_date=row[99],

                        dependent_7_gender=row[100],
                        dependent_7_date_of_birth=row[101],
                        dependent_7_type=row[102],
                        dependent_7_cover_amount=row[103],
                        dependent_7_cover_commencement_date=row[104],

                        dependent_8_gender=row[105],
                        dependent_8_date_of_birth=row[106],
                        dependent_8_type=row[107],
                        dependent_8_cover_amount=row[108],
                        dependent_8_cover_commencement_date=row[109]
                    )

                    funeral_policy.save()
                    new_records.append({"policy_number": policy_number, "status": "created"})

                except IntegrityError as e:
                    # Record creation failed due to integrity error
                    failed_records.append({"policy_number": policy_number, "error": str(e)})
        
        
        except Exception as e:
            return Response({'error': f'Error processing CSV file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        total_records = FuneralPolicy.objects.all().count()

        response_data = {
            'message': 'CSV file uploaded successfully',
            'total_records': total_records,
            'new_records': new_records,
            'updated_records': updated_records,
            'failed_records': failed_records,
        }

        return Response(response_data, status=201)


class IndluLoanDataList(APIView):

    parser_classes = [MultiPartParser]

    def get(self, request, format=None):
        indlu_data_list = IndluLoanData.objects.all()
        serializer = IndluLoanDataSerializer(indlu_data_list, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):

        if 'csv_file' not in request.FILES:
            return Response({'error': 'Please provide a CSV file'}, status=400)
        
        csv_file = request.FILES['csv_file']

        try:
            wb = openpyxl.load_workbook(csv_file)
            worksheet = wb['2023_11_30']

            total_records = 0
            new_records = []
            updated_records = []
            failed_records = []

            for row in worksheet.iter_rows(values_only=True):
                report_data = row[0]
                client_ref = row[1]
                loan_ref_Id = row[2]
                status = row[3]
                close_date = row[4]
                NT = row[5]
                payment_method = row[6]
                merchant = row[7]
                province = row[8]
                disbursment_date = row[9]
                disbursment_month =  row[10]
                application_score = row[11]
                risk_band = row[12]
                debt_book = row[13]
                agency = row[14]
                special_circumstance = row[15]
                sector = row[16]
                job_description = row[17]
                last_payment_date = row[18]
                last_payment_month = row[19]
                gender = row[20]
                employer = row[21]
                loan_amount = row[22]
                int_rate = row[23]
                loan_term = row[24]
                remaining_term = row[25]
                CD_move = row[26]
                CD_current_month = row[27]
                CD_Oct_2023 = row[28]
                CD_Sep_2023 = row[29]
                CD_Aug_2023 = row[30]
                CD_Jul_2023 = row[31]
                CD_Jun_2023 = row[32]
                CD_May_2023 = row[33]
                CD_Apr_2023 = row[34]
                CD_Mar_2023 = row[35]
                CD_Feb_2023 = row[36]
                CD_Jan_2023 = row[37]
                CD_Dec_2023 = row[38]
                CD_Nov_2023 = row[39]
                account_balance_active_charge_off = row[40]

                try:
                    indlu_data_instance = IndluLoanData(
                        report_data=report_data,
                        client_ref=client_ref,
                        loan_ref_Id=loan_ref_Id,
                        status=status,
                        close_date=close_date,
                        NT=NT,
                        payment_method=payment_method,
                        merchant=merchant,
                        province=province,
                        disbursment_date=disbursment_date,
                        disbursment_month=disbursment_month,
                        application_score=application_score,
                        risk_band=risk_band,
                        debt_book=debt_book,
                        agency=agency,
                        special_circumstance=special_circumstance,
                        sector=sector,
                        job_description=job_description,
                        last_payment_date=last_payment_date,
                        last_payment_month=last_payment_month,
                        gender=gender,
                        employer=employer,
                        loan_amount=loan_amount,
                        int_rate=int_rate,
                        loan_term=loan_term,
                        remaining_term=remaining_term,
                        CD_move=CD_move,
                        CD_current_month=CD_current_month,
                        CD_Oct_2023=CD_Oct_2023,
                        CD_Sep_2023=CD_Sep_2023,
                        CD_Aug_2023=CD_Aug_2023,
                        CD_Jul_2023=CD_Jul_2023,
                        CD_Jun_2023=CD_Jun_2023,
                        CD_May_2023=CD_May_2023,
                        CD_Apr_2023=CD_Apr_2023,
                        CD_Mar_2023=CD_Mar_2023,
                        CD_Feb_2023=CD_Feb_2023,
                        CD_Jan_2023=CD_Jan_2023,
                        CD_Dec_2023=CD_Dec_2023,
                        CD_Nov_2023=CD_Nov_2023,
                        account_balance_active_charge_off=account_balance_active_charge_off
                    )
                    indlu_data_instance.save()
                    new_records += 1

                except IntegrityError:
                    failed_records.append(indlu_data_instance)

            total_records = IndluLoanData.objects.all().count()

            return Response({
                'message': 'Indlu Loan data CSV file uploaded successfully',
                'total_records': total_records,
                'new_records': new_records,
                'updated_records': updated_records,
                'failed_records': failed_records
            }, status=201)               

        except Exception as e:
            return Response({'error': f'Error processing CSV file: {e}'}, status=400)

class SmartAdvanceCreditList(APIView):

    parser_classes = [MultiPartParser]
    
    def get(self, request, format=None):
        smart_credit_data_list = SmartAdvanceCredit.objects.all()
        serializer = SmartAdvanceCreditSerializer(smart_credit_data_list, many=True)
        return Response(serializer.data)
    
    def post(self, request, format=None):

        if 'csv_file' not in request.FILES:
            return Response({'error': 'Please provide a CSV file'}, status=status.HTTP_400_BAD_REQUEST)
        
        csv_file = request.FILES['csv_file']

        try:
            wb = openpyxl.load_workbook(csv_file)
            worksheet = wb['Guardrisk Rsa Batches']

            for row in worksheet.iter_rows(values_only=True):
                refId = row[0]
                batch_number = row[1]
                create_date = row[2]
                transmission_date = row[3]
                file_name = row[4]
                loan_ref = row[5]
                policy_status = row[6]
                policy_commencement_date = row[7]
                policy_expiry_date = row[8]
                term_of_policy = row[9]
                policy_status_date = row[10]
                new_policy_indicator = row[11]
                sales_channel = row[12]
                death_premium = row[13]
                ptd_premium = row[14]
                retrenchment_premium = row[15]
                death_original_sum_assured = row[16]
                death_current_sum_assured = row[17]
                PTD_current_sum_assured = row[18]
                retrenchment_current_sum_assured = row[19]
                total_policy_premium_collected = row[20]
                total_policy_premium_payable = row[21]
                original_loan_balance = row[22]
                current_outstanding_balance = row[23]
                installment_amount = row[24]
                principal_surname = row[25]
                principal_first_name = row[26]
                principal_initials = row[27]
                principal_ID = row[28]
                principal_gender = row[29]
                principal_date_of_birth = row[30]
                principal_member_physical_address = row[31]
                principal_member_email_address = row[32]
                principal_telephone_number = row[33]
                postal_code = row[34]
                PTD_original_sum_assured = row[35]
                retrenchment_original_sum_assured = row[36]
                income_group = row[37]
                admin_binder_fees = row[38]
                commission = row[39]
                old_provider = row[40]

                smart_advance_credit_instance = SmartAdvanceCredit(
                    refId = refId,
                    old_provider = old_provider,
                    batch_number = batch_number,
                    create_date = create_date,
                    transmission_date = transmission_date,
                    file_name = file_name,
                    loan_ref = loan_ref,
                    policy_status = policy_status,
                    policy_commencement_date = policy_commencement_date,
                    policy_expiry_date = policy_expiry_date,
                    term_of_policy = term_of_policy,
                    policy_status_date = policy_status_date,
                    new_policy_indicator = new_policy_indicator,
                    sales_channel = sales_channel,
                    death_premium = death_premium,
                    ptd_premium = ptd_premium,
                    retrenchment_premium = retrenchment_premium,
                    death_original_sum_assured = death_original_sum_assured,
                    death_current_sum_assured = death_current_sum_assured,
                    PTD_current_sum_assured = PTD_current_sum_assured,
                    retrenchment_current_sum_assured = retrenchment_current_sum_assured,
                    total_policy_premium_collected = total_policy_premium_collected,
                    total_policy_premium_payable = total_policy_premium_payable,
                    original_loan_balance = original_loan_balance,
                    current_outstanding_balance = current_outstanding_balance,
                    installment_amount = installment_amount,
                    principal_surname = principal_surname,
                    principal_first_name = principal_first_name,
                    principal_initials = principal_initials,
                    principal_ID = principal_ID,
                    principal_gender = principal_gender,
                    principal_date_of_birth = principal_date_of_birth,
                    principal_member_physical_address = principal_member_physical_address,
                    principal_member_email_address = principal_member_email_address,
                    principal_telephone_number = principal_telephone_number,
                    postal_code = postal_code,
                    PTD_original_sum_assured = PTD_original_sum_assured,
                    retrenchment_original_sum_assured = retrenchment_original_sum_assured,
                    income_group = income_group,
                    admin_binder_fees = admin_binder_fees,
                    commission = commission
                )

                smart_advance_credit_instance.save()
            return Response({'message': 'Smart Advance Credit file uploaded successfully'}, status=201)

        except Exception as e:
            return Response({'error': f'Error processing CSV file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

    

def funeral_policy_list(request):
    policies = FuneralPolicy.objects.all()
    return render(request, 'funeral_policy_list.html', {'policies': policies})

def add_funeral_policy(request):
    if request.method == 'POST':
        form = CSVFuneralPolicyUpload(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['csv_file'])
            print('File upload successful')
            return redirect('policy_list')
        else:
            print("Form uploaded is invalid:", form.errors)
    else:
        form = CSVFuneralPolicyUpload()
    return render(request, 'add_funeral_policy.html', {'form': form})



def indlu_data_list(request):
    loan_data_list = IndluLoanData.objects.all()
    return render(request, 'indlu_data_list.html', {'loan_data_list': loan_data_list})

def add_indlu_loan_data(request):
    if request.method == 'POST':
        form = CSVIndluLoanDataUpload(request.POST, request.FILES)
        if form.is_valid():
            handle_indlu_data_upload(request.FILES['csv_file'])
            print('File upload successful')
            return redirect('indlu_list')
        else:
            print("Form uploaded is invalid: ", form.errors)
            return HttpResponse("Form is invalid. Please check the errors.")
    else:
        form = CSVIndluLoanDataUpload()
        return render(request, 'add_indlu_loan_data.html', {'form': form})
    
def smart_credit_data_list(request):
    smart_credit_list = SmartAdvanceCredit.objects.all()
    return render(request, 'smart_credit_data_list.html', {'smart_credit_list': smart_credit_list})

def add_smart_credit_data_list(request):
    if request.method == 'POST':
        form = CSVSmartAdvanceCredit(request.POST, request.FILES)
        if form.is_valid():
            handle_smart_advance_credit_upload(request.FILES['csv_file'])
            print('File upload successful')
            return redirect('smart_credit_list')
        else:
            print('Form uploaded is invalid: ', form.errors)
            return HttpResponse('Form is invalid, inspect for errors')
    else:
        form = CSVSmartAdvanceCredit()
        return render(request, 'add_smart_advance_credit.html',{'form':form})

def handle_indlu_data_upload(file):
    print(type[file])
    wb = openpyxl.load_workbook(file)
    worksheet = wb['2023_11_30']

    for row in worksheet.iter_rows(values_only=True):
        report_data = row[0]
        client_ref = row[1]
        loan_ref_Id = row[2]
        status = row[3]
        close_date = row[4]
        NT = row[5]
        payment_method = row[6]
        merchant = row[7]
        province = row[8]
        disbursment_date = row[9]
        disbursment_month =  row[10]
        application_score = row[11]
        risk_band = row[12]
        debt_book = row[13]
        agency = row[14]
        special_circumstance = row[15]
        sector = row[16]
        job_description = row[17]
        last_payment_date = row[18]
        last_payment_month = row[19]
        gender = row[20]
        employer = row[21]
        loan_amount = row[22]
        int_rate = row[23]
        loan_term = row[24]
        remaining_term = row[25]
        CD_move = row[26]
        CD_current_month = row[27]
        CD_Oct_2023 = row[28]
        CD_Sep_2023 = row[29]
        CD_Aug_2023 = row[30]
        CD_Jul_2023 = row[31]
        CD_Jun_2023 = row[32]
        CD_May_2023 = row[33]
        CD_Apr_2023 = row[34]
        CD_Mar_2023 = row[35]
        CD_Feb_2023 = row[36]
        CD_Jan_2023 = row[37]
        CD_Dec_2023 = row[38]
        CD_Nov_2023 = row[39]
        account_balance_active_charge_off = row[40]

        loan_data = IndluLoanData(
            report_data = report_data,
            client_ref = client_ref,
            loan_ref_Id = loan_ref_Id,
            status = status,
            close_date = close_date,
            NT = NT,
            payment_method = payment_method,
            merchant = merchant,
            province = province,
            disbursment_date = disbursment_date,
            disbursment_month = disbursment_month,
            application_score = application_score,
            risk_band = risk_band,
            debt_book = debt_book,
            agency = agency,
            special_circumstance = special_circumstance,
            sector = sector,
            job_description = job_description,
            last_payment_date = last_payment_date,
            last_payment_month = last_payment_month,
            gender = gender,
            employer = employer,
            loan_amount = loan_amount,
            int_rate = int_rate,
            loan_term = loan_term,
            remaining_term = remaining_term,
            CD_move = CD_move,
            CD_current_month = CD_current_month,
            CD_Oct_2023 = CD_Oct_2023,
            CD_Sep_2023 = CD_Sep_2023,
            CD_Aug_2023 = CD_Aug_2023,
            CD_Jul_2023 = CD_Jul_2023,
            CD_Jun_2023 = CD_Jun_2023,
            CD_May_2023 = CD_May_2023,
            CD_Apr_2023 = CD_Apr_2023,
            CD_Mar_2023 = CD_Mar_2023,
            CD_Feb_2023 = CD_Feb_2023,
            CD_Jan_2023 = CD_Jan_2023,
            CD_Dec_2023 = CD_Dec_2023,
            CD_Nov_2023 = CD_Nov_2023,
            account_balance_active_charge_off = account_balance_active_charge_off
        )

        try:
            loan_data.save()
        except Exception as e:
            print(f"Error saving IndluLoan data instance :{e}")  


def handle_uploaded_file(file):
    print(type[file])
    wb = openpyxl.load_workbook(file)
    worksheet = wb["Funeral Policies"]

    for row in worksheet.iter_rows(values_only=True):
        policy_search = row[0]
        ID_search = row[1]
        time_stamp = row[2]
        report_period_start = row[3]
        report_period_end = row[4]
        administrator = row[5]
        insurer_name = row[6]
        client_identifier = row[7]
        division_identifier = row[8]
        sub_scheme_name = row[9]
        policy_number = row[10]
        product_name = row[11]
        product_option = row[12]
        policy_commencement_date = row[13]
        policy_expiry_date = row[14]
        term_of_policy = row[15]
        policy_status = row[16]
        policy_status_date = row[17]
        new_policy_indicator = row[18]
        sales_channel = row[19]
        cancelled_by = row[20]
        death_indicator = row[21]
        premium_frequency = row[22]
        premium_type = row[23]
        death_original_sum_insured = row[24]
        death_cover = row[25]
        death_current_sum_insured = row[26]
        reinsurer_name = row[27]
        death_current_ri_sum_insured = row[28]
        death_ri_premium = row[29]
        death_ri_percent = row[30]
        total_premium_collected = row[31]
        total_premium_payable = row[32]
        total_premium_subsidy = row[33]
        total_reinsurance_premium = row[34]
        total_reinsurance_premium_payable = row[35]
        total_financial_reinsurance_cashflows = row[36]
        total_financial_reinsurance_payable = row[37]
        commission_frequency = row[38]
        commission = row[39]
        binder_fees = row[40]
        outsourcing_fees = row[41]
        marketing_fees = row[42]
        management_fees = row[43]
        claims_handling_fees = row[44]
        total_gross_claim_amount = row[45]
        gross_claim_paid = row[46]
        reinsurance_recoveries = row[47]
        principal_surname = row[48]
        principal_firstname = row[49]
        principal_initials = row[50]
        principal_id = row[51]
        principal_gender = row[52]
        principal_date_of_birth = row[53]
        principal_physical_address = row[54]
        principal_postal_code = row[55]
        principal_email = row[56]
        income_group = row[57]
        spouse_indicator = row[58]
        number_adult_dependants = row[59]
        number_child_dependants = row[60]
        number_extended_family = row[61]
        spouse_surname = row[62]
        spouse_firstname = row[63]
        spouse_initials = row[64]
        spouse_id = row[65]
        spouse_gender = row[66]
        spouse_date_of_birth = row[67]
        spouse_cover_amount = row[68]
        spouse_cover_commencement_date = row[69]
        dependent_1_gender = row[70]
        dependent_1_date_of_birth = row[80]
        dependent_1_type = row[81]
        dependent_1_cover_amount = row[82]
        dependent_1_cover_commencement_date = row[83]
        dependent_2_gender = row[84]
        dependent_2_date_of_birth = row[85]
        dependent_2_type = row[86]
        dependent_2_cover_amount = row[87]
        dependent_2_cover_commencement_date = row[88]
        dependent_3_gender = row[89]
        dependent_3_date_of_birth = row[90]
        dependent_3_type = row[91]
        dependent_3_cover_amount = row[92]
        dependent_3_cover_commencement_date = row[93]
        dependent_4_gender = row[94]
        dependent_4_date_of_birth = row[95]
        dependent_4_type = row[96]
        dependent_4_cover_amount = row[97]
        dependent_4_cover_commencement_date = row[98]
        dependent_5_gender = row[99]
        dependent_5_date_of_birth = row[100]
        dependent_5_type = row[101]
        dependent_5_cover_amount = row[102]
        dependent_5_cover_commencement_date = row[103]
        dependent_6_gender = row[104]
        dependent_6_date_of_birth = row[105]
        dependent_6_type = row[106]
        dependent_6_cover_amount = row[107]
        dependent_6_cover_commencement_date = row[108]
        dependent_7_gender = row[109]
        dependent_7_date_of_birth = row[110]
        dependent_7_type = row[111]
        dependent_7_cover_amount = row[112]
        dependent_7_cover_commencement_date = row[113]
        dependent_8_gender = row[114]
        dependent_8_date_of_birth = row[115]
        dependent_8_type = row[116]
        dependent_8_cover_amount = row[117]
        dependent_8_cover_commencement_date = row[118]
       
        funeral_policy = FuneralPolicy(
            policy_search = policy_search,
            ID_search = ID_search,
            time_stamp = time_stamp,
            report_period_start = report_period_start,
            report_period_end = report_period_end,
            administrator = administrator,
            insurer_name = insurer_name,
            client_identifier = client_identifier,
            division_identifier = division_identifier,
            sub_scheme_name = sub_scheme_name,
            policy_number = policy_number,
            product_name = product_name,
            product_option = product_option,
            policy_commencement_date = policy_commencement_date,
            policy_expiry_date = policy_expiry_date,
            term_of_policy = term_of_policy,
            policy_status = policy_status,
            policy_status_date = policy_status_date,
            new_policy_indicator = new_policy_indicator,
            sales_channel = sales_channel,
            cancelled_by = cancelled_by,
            death_indicator = death_indicator,
            premium_frequency = premium_frequency,
            premium_type = premium_type,
            death_original_sum_insured = death_original_sum_insured,
            death_cover = death_cover,
            death_current_sum_insured = death_current_sum_insured,
            reinsurer_name = reinsurer_name,
            death_current_ri_sum_insured = death_current_ri_sum_insured,
            death_ri_premium = death_ri_premium,
            death_ri_percent = death_ri_percent,
            total_premium_collected = total_premium_collected,
            total_premium_payable = total_premium_payable,
            total_premium_subsidy = total_premium_subsidy,
            total_reinsurance_premium = total_reinsurance_premium,
            total_reinsurance_premium_payable = total_reinsurance_premium_payable,
            total_financial_reinsurance_cashflows = total_financial_reinsurance_cashflows,
            total_financial_reinsurance_payable = total_financial_reinsurance_payable,
            commission_frequency = commission_frequency,
            commission = commission,
            binder_fees = binder_fees,
            outsourcing_fees = outsourcing_fees,
            marketing_fees = marketing_fees,
            management_fees = management_fees,
            claims_handling_fees = claims_handling_fees,
            total_gross_claim_amount = total_gross_claim_amount,
            gross_claim_paid = gross_claim_paid,
            reinsurance_recoveries = reinsurance_recoveries,
            principal_surname = principal_surname,
            principal_firstname = principal_firstname,
            principal_initials = principal_initials,
            principal_id = principal_id,
            principal_gender = principal_gender,
            principal_date_of_birth =  principal_date_of_birth,
            principal_physical_address = principal_physical_address,
            principal_postal_code = principal_postal_code,
            principal_email = principal_email,
            income_group = income_group,
            spouse_indicator = spouse_indicator,
            number_adult_dependants = number_adult_dependants,
            number_child_dependants = number_child_dependants,
            number_extended_family = number_extended_family,
            spouse_surname = spouse_surname,
            spouse_firstname = spouse_firstname,
            spouse_initials = spouse_initials,
            spouse_id = spouse_id,
            spouse_gender = spouse_gender,
            spouse_date_of_birth = spouse_date_of_birth,
            spouse_cover_amount = spouse_cover_amount,
            spouse_cover_commencement_date = spouse_cover_commencement_date,

            dependent_1_gender = dependent_1_gender,
            dependent_1_date_of_birth = dependent_1_date_of_birth,
            dependent_1_type =  dependent_1_type,
            dependent_1_cover_amount = dependent_1_cover_amount,
            dependent_1_cover_commencement_date = dependent_1_cover_commencement_date,    

            dependent_2_gender = dependent_2_gender,
            dependent_2_date_of_birth = dependent_2_date_of_birth,
            dependent_2_type =  dependent_2_type,
            dependent_2_cover_amount = dependent_2_cover_amount,
            dependent_2_cover_commencement_date = dependent_2_cover_commencement_date,

            dependent_3_gender = dependent_3_gender,
            dependent_3_date_of_birth = dependent_3_date_of_birth,
            dependent_3_type =  dependent_3_type,
            dependent_3_cover_amount = dependent_3_cover_amount,
            dependent_3_cover_commencement_date = dependent_3_cover_commencement_date,

            dependent_4_gender = dependent_4_gender,
            dependent_4_date_of_birth = dependent_4_date_of_birth,
            dependent_4_type =  dependent_4_type,
            dependent_4_cover_amount = dependent_4_cover_amount,
            dependent_4_cover_commencement_date = dependent_4_cover_commencement_date,

            dependent_5_gender = dependent_5_gender,
            dependent_5_date_of_birth = dependent_5_date_of_birth,
            dependent_5_type =  dependent_5_type,
            dependent_5_cover_amount = dependent_5_cover_amount,
            dependent_5_cover_commencement_date = dependent_5_cover_commencement_date,

            dependent_6_gender = dependent_6_gender,
            dependent_6_date_of_birth = dependent_6_date_of_birth,
            dependent_6_type =  dependent_6_type,
            dependent_6_cover_amount = dependent_6_cover_amount,
            dependent_6_cover_commencement_date = dependent_6_cover_commencement_date,

            dependent_7_gender = dependent_7_gender,
            dependent_7_date_of_birth = dependent_7_date_of_birth,
            dependent_7_type =  dependent_7_type,
            dependent_7_cover_amount = dependent_7_cover_amount,
            dependent_7_cover_commencement_date = dependent_7_cover_commencement_date,  

            dependent_8_gender = dependent_8_gender,
            dependent_8_date_of_birth = dependent_8_date_of_birth,
            dependent_8_type =  dependent_8_type,
            dependent_8_cover_amount = dependent_8_cover_amount,
            dependent_8_cover_commencement_date = dependent_8_cover_commencement_date,      
        )

        try:
            funeral_policy.save()
        except Exception as e:
            print(f"Error saving funeral policy instance :{e}")


def handle_smart_advance_credit_upload(file):
    print(type[file])
    wb = openpyxl.load_workbook(file)
    worksheet = wb['Guardrisk Rsa Batches']

    for row in worksheet.iter_rows(values_only=True):
        refId = row[0]
        batch_number = row[1]
        create_date = row[2]
        transmission_date = row[3]
        file_name = row[4]
        loan_ref = row[5]
        policy_status = row[6]
        policy_commencement_date = row[7]
        policy_expiry_date = row[8]
        term_of_policy = row[9]
        policy_status_date = row[10]
        new_policy_indicator = row[11]
        sales_channel = row[12]
        death_premium = row[13]
        ptd_premium = row[14]
        retrenchment_premium = row[15]
        death_original_sum_assured = row[16]
        death_current_sum_assured = row[17]
        PTD_current_sum_assured = row[18]
        retrenchment_current_sum_assured = row[19]
        total_policy_premium_collected = row[20]
        total_policy_premium_payable = row[21]
        original_loan_balance = row[22]
        current_outstanding_balance = row[23]
        installment_amount = row[24]
        principal_surname = row[25]
        principal_first_name = row[26]
        principal_initials = row[27]
        principal_ID = row[28]
        principal_gender = row[29]
        principal_date_of_birth = row[30]
        principal_member_physical_address = row[31]
        principal_member_email_address = row[32]
        principal_telephone_number = row[33]
        postal_code = row[34]
        PTD_original_sum_assured = row[35]
        retrenchment_original_sum_assured = row[36]
        income_group = row[37]
        admin_binder_fees = row[38]
        commission = row[39]
        old_provider = row[40]

        smart_credit_data_instance = SmartAdvanceCredit(
            refId = refId,
            old_provider = old_provider,
            batch_number = batch_number,
            create_date = create_date,
            transmission_date = transmission_date,
            file_name = file_name,
            loan_ref = loan_ref,
            policy_status = policy_status,
            policy_commencement_date = policy_commencement_date,
            policy_expiry_date = policy_expiry_date,
            term_of_policy = term_of_policy,
            policy_status_date = policy_status_date,
            new_policy_indicator = new_policy_indicator,
            sales_channel = sales_channel,
            death_premium = death_premium,
            ptd_premium = ptd_premium,
            retrenchment_premium = retrenchment_premium,
            death_original_sum_assured = death_original_sum_assured,
            death_current_sum_assured = death_current_sum_assured,
            PTD_current_sum_assured = PTD_current_sum_assured,
            retrenchment_current_sum_assured = retrenchment_current_sum_assured,
            total_policy_premium_collected = total_policy_premium_collected,
            total_policy_premium_payable = total_policy_premium_payable,
            original_loan_balance = original_loan_balance,
            current_outstanding_balance = current_outstanding_balance,
            installment_amount = installment_amount,
            principal_surname = principal_surname,
            principal_first_name = principal_first_name,
            principal_initials = principal_initials,
            principal_ID = principal_ID,
            principal_gender = principal_gender,
            principal_date_of_birth = principal_date_of_birth,
            principal_member_physical_address = principal_member_physical_address,
            principal_member_email_address = principal_member_email_address,
            principal_telephone_number = principal_telephone_number,
            postal_code = postal_code,
            PTD_original_sum_assured = PTD_original_sum_assured,
            retrenchment_original_sum_assured = retrenchment_original_sum_assured,
            income_group = income_group,
            admin_binder_fees = admin_binder_fees,
            commission = commission
        )

        try:
            smart_credit_data_instance.save()
        except Exception as e:
            print(f"Error saving Smart credit policy instance : {e}")